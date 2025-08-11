LETTER_MAP = {
    'A': [
        [0,1,1,1,0],
        [1,0,0,0,1],
        [1,1,1,1,1],
        [1,0,0,0,1],
        [1,0,0,0,1],
    ],
    'B': [
        [1,1,1,1,0],
        [1,0,0,0,1],
        [1,1,1,1,0],
        [1,0,0,0,1],
        [1,1,1,1,0],
    ],
    'C': [
        [0,1,1,1,1],
        [1,0,0,0,0],
        [1,0,0,0,0],
        [1,0,0,0,0],
        [0,1,1,1,1],
    ],
    'D': [
        [1,1,1,1,0],
        [1,0,0,0,1],
        [1,0,0,0,1],
        [1,0,0,0,1],
        [1,1,1,1,0],
    ],
    'E': [
        [1,1,1,1,1],
        [1,0,0,0,0],
        [1,1,1,1,0],
        [1,0,0,0,0],
        [1,1,1,1,1],
    ],
    'F': [
        [1,1,1,1,1],
        [1,0,0,0,0],
        [1,1,1,1,0],
        [1,0,0,0,0],
        [1,0,0,0,0],
    ],
    'G': [
        [0,1,1,1,1],
        [1,0,0,0,0],
        [1,0,0,1,1],
        [1,0,0,0,1],
        [0,1,1,1,1],
    ],
    'H': [
        [1,0,0,0,1],
        [1,0,0,0,1],
        [1,1,1,1,1],
        [1,0,0,0,1],
        [1,0,0,0,1],
    ],
    'I': [
        [1,1,1,1,1],
        [0,0,1,0,0],
        [0,0,1,0,0],
        [0,0,1,0,0],
        [1,1,1,1,1],
    ],
    'J': [
        [0,0,0,1,1],
        [0,0,0,0,1],
        [0,0,0,0,1],
        [1,0,0,0,1],
        [0,1,1,1,0],
    ],
    'K': [
        [1,0,0,0,1],
        [1,0,0,1,0],
        [1,1,1,0,0],
        [1,0,0,1,0],
        [1,0,0,0,1],
    ],
    'L': [
        [1,0,0,0,0],
        [1,0,0,0,0],
        [1,0,0,0,0],
        [1,0,0,0,0],
        [1,1,1,1,1],
    ],
    'M': [
        [1,0,0,0,1],
        [1,1,0,1,1],
        [1,0,1,0,1],
        [1,0,0,0,1],
        [1,0,0,0,1],
    ],
    'N': [
        [1,0,0,0,1],
        [1,1,0,0,1],
        [1,0,1,0,1],
        [1,0,0,1,1],
        [1,0,0,0,1],
    ],
    'O': [
        [0,1,1,1,0],
        [1,0,0,0,1],
        [1,0,0,0,1],
        [1,0,0,0,1],
        [0,1,1,1,0],
    ],
    'P': [
        [1,1,1,1,0],
        [1,0,0,0,1],
        [1,1,1,1,0],
        [1,0,0,0,0],
        [1,0,0,0,0],
    ],
    'Q': [
        [0,1,1,1,0],
        [1,0,0,0,1],
        [1,0,0,0,1],
        [1,0,1,0,1],
        [0,1,1,1,1],
    ],
    'R': [
        [1,1,1,1,0],
        [1,0,0,0,1],
        [1,1,1,1,0],
        [1,0,0,1,0],
        [1,0,0,0,1],
    ],
    'S': [
        [0,1,1,1,1],
        [1,0,0,0,0],
        [0,1,1,1,0],
        [0,0,0,0,1],
        [1,1,1,1,0],
    ],
    'T': [
        [1,1,1,1,1],
        [0,0,1,0,0],
        [0,0,1,0,0],
        [0,0,1,0,0],
        [0,0,1,0,0],
    ],
    'U': [
        [1,0,0,0,1],
        [1,0,0,0,1],
        [1,0,0,0,1],
        [1,0,0,0,1],
        [0,1,1,1,0],
    ],
    'V': [
        [1,0,0,0,1],
        [1,0,0,0,1],
        [0,1,0,1,0],
        [0,1,0,1,0],
        [0,0,1,0,0],
    ],
    'W': [
        [1,0,0,0,1],
        [1,0,0,0,1],
        [1,0,1,0,1],
        [1,1,0,1,1],
        [1,0,0,0,1],
    ],
    'X': [
        [1,0,0,0,1],
        [0,1,0,1,0],
        [0,0,1,0,0],
        [0,1,0,1,0],
        [1,0,0,0,1],
    ],
    'Y': [
        [1,0,0,0,1],
        [0,1,0,1,0],
        [0,0,1,0,0],
        [0,0,1,0,0],
        [0,0,1,0,0],
    ],
    'Z': [
        [1,1,1,1,1],
        [0,0,0,1,0],
        [0,0,1,0,0],
        [0,1,0,0,0],
        [1,1,1,1,1],
    ],
}

import random
import string

from openpyxl import Workbook
from openpyxl.styles import PatternFill

import tkinter as tk
from tkinter.font import Font
from tkinter import ttk, messagebox, filedialog

# Generates a random string with letters, digits, and symbols
def random_string(length=4):
    characters = string.ascii_letters + string.digits + "!@#$%^&*()_+-"
    return ''.join(random.choices(characters, k=length))

# Creates a matrix (2D list) of random strings
def create_matrix(rows, cols):
    return [[random_string() for _ in range(cols)] for _ in range(rows)]

# Extracts a password based on the given letter shape (LETTER_MAP) and character index
def extract_password(matrix, start_row, start_col, letter, char_index):
    shape = LETTER_MAP[letter]
    password = ''
    for i in range(len(shape)):
        for j in range(len(shape[0])):
            if shape[i][j] == 1:
                if 0 <= start_row + i < len(matrix) and 0 <= start_col + j < len(matrix[0]):
                    cell = matrix[start_row + i][start_col + j]
                    if len(cell) > char_index:
                        password += cell[char_index]
    return password

# Debug utility: print the matrix to console
def print_matrix(matrix):
    for row in matrix:
        print(' '.join(row))
    print()

# Save the matrix to an Excel file and highlight the shape for the selected letter
def save_matrix_to_excel(matrix, start_row, start_col, letter, filename='matrix_output.xlsx'):
    wb = Workbook()
    ws = wb.active
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    shape = LETTER_MAP.get(letter)
    for i, row in enumerate(matrix):
        for j, cell_value in enumerate(row):
            cell = ws.cell(row=i+1, column=j+1, value=cell_value)
            if shape:
                map_i = i - start_row
                map_j = j - start_col
                if 0 <= map_i < len(shape) and 0 <= map_j < len(shape[0]):
                    if shape[map_i][map_j] == 1:
                        cell.fill = yellow_fill
    wb.save(filename)
    print(f"\nExcel file has been saved as '{filename}'.")

# Main GUI application class
class PasswordGeneratorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Password Matrix Generator")
        self.root.geometry("800x600")
        self.style = ttk.Style()
        self.setup_style()
        self.create_widgets()
    
    # Setup UI styling for a consistent look
    def setup_style(self):
        self.style.theme_use('clam')
        self.style.configure('TFrame', background='#f0f0f0')
        self.style.configure('TLabel', background='#f0f0f0', font=('Helvetica', 10))
        self.style.configure('TButton', font=('Helvetica', 10), padding=5)
        self.style.configure('Header.TLabel', font=('Helvetica', 14, 'bold'))
        self.style.configure('Output.TLabel', font=('Courier', 12))
        self.style.map('TButton',
                      foreground=[('active', 'black'), ('!active', 'black')],
                      background=[('active', '#d9d9d9'), ('!active', '#e6e6e6')])

    # Create and lay out all UI components
    def create_widgets(self):
        # Main frame
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Header label
        header = ttk.Label(main_frame, text="Password Matrix Generator", style='Header.TLabel')
        header.pack(pady=(0, 20))
        
        # Settings frame
        settings_frame = ttk.LabelFrame(main_frame, text="Settings", padding=10)
        settings_frame.pack(fill=tk.X, pady=10)
        
        # Letter selection
        ttk.Label(settings_frame, text="Select Letter:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        self.letter_var = tk.StringVar()
        self.letter_dropdown = ttk.Combobox(settings_frame, textvariable=self.letter_var, 
                                        values=list(LETTER_MAP.keys()), state="readonly")
        self.letter_dropdown.grid(row=0, column=1, sticky=tk.W, padx=5, pady=5)
        self.letter_dropdown.current(0)
        
        # Character index selection
        ttk.Label(settings_frame, text="Character Index:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=5)
        self.index_var = tk.StringVar()
        self.index_dropdown = ttk.Combobox(settings_frame, textvariable=self.index_var, 
                                        values=[1, 2, 3, 4], state="readonly")
        self.index_dropdown.grid(row=1, column=1, sticky=tk.W, padx=5, pady=5)
        self.index_dropdown.current(0)
        
        # Buttons frame
        buttons_frame = ttk.Frame(main_frame)
        buttons_frame.pack(fill=tk.X, pady=10)
        
        generate_btn = ttk.Button(buttons_frame, text="Generate Password", command=self.generate_password)
        generate_btn.pack(side=tk.LEFT, padx=5)
        
        save_btn = ttk.Button(buttons_frame, text="Save to Excel", command=self.save_to_excel)
        save_btn.pack(side=tk.LEFT, padx=5)
        
        # Output frame
        output_frame = ttk.LabelFrame(main_frame, text="Output", padding=10)
        output_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        self.password_var = tk.StringVar()
        password_label = ttk.Label(output_frame, text="Password:", style='Header.TLabel')
        password_label.pack(anchor=tk.W)
        
        self.password_output = ttk.Label(output_frame, textvariable=self.password_var, 
                                    style='Output.TLabel', background='white', 
                                    relief=tk.SUNKEN, padding=5, anchor=tk.CENTER)
        self.password_output.pack(fill=tk.X, pady=5)
        
        # Matrix display
        self.matrix_text = tk.Text(output_frame, height=15, width=50, font=('Courier', 8))
        scrollbar = ttk.Scrollbar(output_frame, orient=tk.VERTICAL, command=self.matrix_text.yview)
        self.matrix_text.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.matrix_text.pack(fill=tk.BOTH, expand=True)
        
        # Generate the matrix on app start
        self.matrix = None
        self.generate_matrix()
    
    # Creates and stores a new matrix
    def generate_matrix(self):
        size = [10, 10]
        rows, cols = int(size[0]), int(size[1])
        self.matrix = create_matrix(rows, cols)
        self.display_matrix()
        
    # Displays the matrix in the text box
    def display_matrix(self):
        self.matrix_text.delete(1.0, tk.END)
        for row in self.matrix:
            self.matrix_text.insert(tk.END, ' '.join(row) + '\n')
    
    # Handles password generation when button is clicked
    def generate_password(self):
        chosen_letter = self.letter_var.get()
        try:
            char_index = int(self.index_var.get()) - 1
            if not (0 <= char_index <= 3):
                raise ValueError("Character index must be between 1 and 4.")
        except ValueError as e:
            messagebox.showerror("Error", f"Invalid input: {e}")
            return
            
        start_row, start_col = 2, 2 # Starting position in the matrix
        if chosen_letter in LETTER_MAP:
            password = extract_password(self.matrix, start_row, start_col, chosen_letter, char_index)
            self.password_var.set(password)
            self.display_matrix()
        else:
            messagebox.showerror("Error", "There is no defined pattern for this letter.")

    # Handles saving the matrix to an Excel file        
    def save_to_excel(self):
        chosen_letter = self.letter_var.get()
        if chosen_letter not in LETTER_MAP:
            messagebox.showerror("Error", "Please select a valid letter first.")
            return
            
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Save Excel File"
        )
        
        if file_path:
            start_row, start_col = 2, 2
            save_matrix_to_excel(self.matrix, start_row, start_col, chosen_letter, file_path)
            messagebox.showinfo("Success", f"Excel file has been saved as '{file_path}'")

if __name__ == "__main__":
    root = tk.Tk()
    app = PasswordGeneratorApp(root)
    root.mainloop()